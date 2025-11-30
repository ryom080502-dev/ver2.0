import os
import json
import time
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
import streamlit as st
import google.generativeai as genai
from dotenv import load_dotenv

# --- 設定 ---
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY") 
MODEL_NAME = "gemini-2.5-flash"
TEMPLATE_FILE = "template.xlsx"

# ▼▼▼ 合言葉の設定 ▼▼▼
LOGIN_PASSWORD = "fujishima8888" 
# ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲

# --- ページ設定 ---
st.set_page_config(page_title="経費精算AI", layout="wide")

# ▼▼▼ CSSスタイル ▼▼▼
st.markdown("""
    <style>
    [data-testid="stFileUploaderDropzoneInstructions"] > div > span {display: none;}
    [data-testid="stFileUploaderDropzoneInstructions"] > div::after {
        content: "ファイルをドラッグまたは選択"; font-weight: bold; font-size: 1rem;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] > div > small {display: none;}
    [data-testid="stFileUploaderDropzoneInstructions"] > div::before {
        content: "上限 200MB / PDFのみ"; font-size: 0.8rem; display: block; margin-bottom: 5px;
    }
    [data-testid="stMetric"] {
        background-color: #f0f2f6; padding: 15px; border-radius: 10px; border: 1px solid #e0e0e0;
    }
    @media (prefers-color-scheme: dark) {
        [data-testid="stMetric"] { background-color: #262730; border: 1px solid #41444e; }
    }
    </style>
""", unsafe_allow_html=True)

# --- 認証機能 ---
def check_password():
    if 'authenticated' not in st.session_state:
        st.session_state['authenticated'] = False
    if st.session_state['authenticated']:
        return True

    st.title("🔒 ログイン")
    password = st.text_input("パスワードを入力してください", type="password")
    if st.button("ログイン"):
        if password == LOGIN_PASSWORD:
            st.session_state['authenticated'] = True
            st.rerun()
        else:
            st.error("パスワードが違います")
    return False

# --- 結合セル対応の書き込み関数 ---
def smart_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    else:
        cell.value = value

# --- メインロジック関数 ---
def analyze_and_create_excel(uploaded_file, template_path, output_excel_path):
    api_key_to_use = API_KEY
    if not api_key_to_use and "GOOGLE_API_KEY" in st.secrets:
        api_key_to_use = st.secrets["GOOGLE_API_KEY"]

    if not api_key_to_use:
        st.error("APIキー設定エラー")
        return None

    genai.configure(api_key=api_key_to_use)
    
    model = genai.GenerativeModel(
        model_name=MODEL_NAME,
        generation_config={"temperature": 0, "response_mime_type": "application/json"},
        system_instruction="""
        あなたは優秀な経理担当アシスタントです。
        ユーザーからアップロードされるPDFファイルは、複数のレシートや領収書を連続でスキャンしたデータです。
        以下のルールに従って、画像内の情報を解析し、正確なJSONデータとして出力してください。
        
        ### 抽出・判定ルール
        1. **日付 (date):** `YYYY/MM/DD` 形式。不明な場合は `null`。
        2. **店名 (store_name):** 店舗名。不明な場合は `null`。
        3. **インボイス登録番号 (invoice_number):** `T`から始まる13桁の番号があれば抽出。なければ `null`。
        4. **金額の内訳:**
           - **amount_8_percent:** 税率8%（軽減税率・食品など）の対象金額（税込）。
           - **amount_10_percent:** 税率10%の対象金額（税込）。
           - **amount_non_invoice:** インボイス登録番号がない、または区分不明な金額。
        
        ### エラーハンドリング
        - 読み取れない箇所がある場合でも、読み取れた項目は必ず出力する。
        - 全く読めない場合は `status` を `error` とする。
        
        ### 出力フォーマット (JSON List)
        [{"status": "success", "date": "...", "store_name": "...", "invoice_number": "T...", "amount_8_percent": 0, "amount_10_percent": 0, "amount_non_invoice": 0, "error_message": null}]
        """
    )

    try:
        temp_pdf_path = "temp_input.pdf"
        with open(temp_pdf_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        sample_file = genai.upload_file(path=temp_pdf_path, display_name="User Upload PDF")
        
        with st.spinner(' レシートを読み込んでいます (数分かかる場合があります)'):
            while sample_file.state.name == "PROCESSING":
                time.sleep(1)
                sample_file = genai.get_file(sample_file.name)

            if sample_file.state.name == "FAILED":
                st.error("Google側での処理に失敗しました")
                return None

            response = model.generate_content([sample_file, "このPDFの全ページのレシート情報を抽出してください。"])
            receipt_data = json.loads(response.text)

        # 日付ソート
        receipt_data.sort(key=lambda x: x.get("date") if x.get("date") else "9999/99/99")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active 
        
        # ▼▼▼ 修正箇所: 行番号の計算ロジックを変更 ▼▼▼
        # 9行目～29行目 = 21行分 (インデックス0～20)
        # 41行目～      = それ以降 (インデックス21～)
        
        for i, item in enumerate(receipt_data):
            if i <= 20:
                # 1ページ目 (0~20件目) -> 9行目スタート
                row_num = 9 + i
            else:
                # 2ページ目 (21件目以降) -> 41行目スタート
                # 例: i=21のとき、41 + (21-21) = 41行目
                row_num = 41 + (i - 21)

            # --- ここから書き込み処理 ---
            if item.get("date"): 
                smart_write(ws, row_num, 2, item["date"])
            
            if item.get("store_name"): 
                smart_write(ws, row_num, 5, item["store_name"]) 
            
            amt_8 = item.get("amount_8_percent") or 0
            amt_10 = item.get("amount_10_percent") or 0
            amt_other = item.get("amount_non_invoice") or 0

            total_8_zone = amt_8 + amt_other
            if total_8_zone > 0: 
                smart_write(ws, row_num, 16, total_8_zone) 
            
            if amt_10 > 0: 
                smart_write(ws, row_num, 19, amt_10)

        wb.save(output_excel_path)
        return receipt_data

    except Exception as e:
        st.error(f"システムエラー: {e}")
        return None

# --- メイン処理 ---
if check_password():
    st.title("🧾 経費精算 自動入力アプリ")
    st.markdown("---")

    col1, col2 = st.columns([1, 2.5])

    with col1:
        st.subheader("📂 1. ファイル選択")
        uploaded_file = st.file_uploader("PDFアップロード", type=["pdf"])
        
        if uploaded_file is not None:
            st.success("ファイル選択済み")
            st.write("")
            st.subheader("🚀 2. 実行")
            if st.button("読み取りを開始", type="primary", use_container_width=True):
                temp_excel_path = "result_download.xlsx"
                if os.path.exists(TEMPLATE_FILE):
                    result_data = analyze_and_create_excel(uploaded_file, TEMPLATE_FILE, temp_excel_path)
                    if result_data:
                        st.session_state['result_data'] = result_data
                        st.session_state['excel_ready'] = True
                else:
                    st.error(f"テンプレート ({TEMPLATE_FILE}) が見つかりません。")
            
            if 'excel_ready' in st.session_state:
                st.write("")
                st.write("---")
                with open("result_download.xlsx", "rb") as f:
                    st.download_button(
                        label="📥 経費精算書をダウンロード",
                        data=f,
                        file_name=f"経費精算_{os.path.basename('result_download.xlsx')}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="secondary",
                        use_container_width=True
                    )

    with col2:
        st.subheader("📊 解析結果ダッシュボード")
        
        if 'result_data' in st.session_state:
            data = st.session_state['result_data']
            
            total_10 = sum([d.get("amount_10_percent", 0) for d in data])
            total_8 = sum([d.get("amount_8_percent", 0) for d in data])
            total_other = sum([d.get("amount_non_invoice", 0) for d in data])
            count = len(data)

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("読取枚数", f"{count} 枚")
            m2.metric("10%対象", f"¥{total_10:,}")
            m3.metric("8%対象", f"¥{total_8:,}")
            m4.metric("対象外・不明", f"¥{total_other:,}")

            st.write("")

            df = pd.DataFrame(data)
            df["total_amount"] = df.apply(lambda x: x.get("amount_10_percent", 0) + x.get("amount_8_percent", 0) + x.get("amount_non_invoice", 0), axis=1)
            
            def format_invoice(row):
                num = row.get("invoice_number")
                if num and str(num).startswith("T") and len(str(num)) >= 13:
                    return f"✅ 適合 ({num})"
                else:
                    return "➖ 非適合"
            
            df["invoice_status"] = df.apply(format_invoice, axis=1)

            df_display = df[[
                "date", "store_name", "total_amount", "invoice_status", 
                "amount_10_percent", "amount_8_percent", "amount_non_invoice"
            ]].rename(columns={
                "date": "日付",
                "store_name": "店舗名",
                "total_amount": "支払総額",
                "invoice_status": "インボイス",
                "amount_10_percent": "10%対象",
                "amount_8_percent": "8%対象",
                "amount_non_invoice": "対象外/不明"
            })

            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "支払総額": st.column_config.NumberColumn(format="¥%d"),
                    "10%対象": st.column_config.NumberColumn(format="¥%d"),
                    "8%対象": st.column_config.NumberColumn(format="¥%d"),
                    "対象外/不明": st.column_config.NumberColumn(format="¥%d"),
                    "インボイス": st.column_config.TextColumn(width="medium"),
                }
            )

        else:
            st.info("👈 左側のボタンを押して読み取りを開始してください。")
            cols = st.columns(4)
            for c in cols: c.metric("---", "---")
            st.dataframe(pd.DataFrame({"日付":[], "店舗名":[], "支払総額":[]}), use_container_width=True)