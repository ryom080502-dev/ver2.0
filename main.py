import os
import json
import time
import openpyxl
import google.generativeai as genai
from dotenv import load_dotenv

# 1. 環境変数の読み込み
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

if not API_KEY:
    raise ValueError("APIキーが設定されていません。.envファイルを確認してください。")

genai.configure(api_key=API_KEY)

# ▼ モデル設定
MODEL_NAME = "gemini-2.5-flash"

# 2. Geminiの設定
generation_config = {
    "temperature": 0,
    "response_mime_type": "application/json",
}

model = genai.GenerativeModel(
    model_name=MODEL_NAME,
    generation_config=generation_config,
    system_instruction="""
    あなたは優秀な経理担当アシスタントです。
    ユーザーからアップロードされるPDFファイルは、複数のレシートや領収書を連続でスキャンしたデータです。
    以下のルールに従って、画像内の情報を解析し、正確なJSONデータとして出力してください。

    ### 抽出・判定ルール
    1. **日付 (date):** `YYYY/MM/DD` 形式。不明な場合は `null`。
    2. **店名 (store_name):** 店舗名。不明な場合は `null`。
    3. **金額の内訳:**
       - **amount_8_percent:** 税率8%（軽減税率・食品など）の対象金額（税込）。
       - **amount_10_percent:** 税率10%の対象金額（税込）。
       - **amount_non_invoice:** インボイス登録番号がない、または区分不明な金額。
    
    ### エラーハンドリング
    - 読み取れない箇所がある場合でも、読み取れた項目（日付だけ、店名だけ等）は必ず出力する。
    - 全く読めない場合は `status` を `error` とする。

    ### 出力フォーマット (JSON List)
    [
      {
        "status": "success",
        "date": "2024/11/29",
        "store_name": "店舗名",
        "amount_8_percent": 500,
        "amount_10_percent": 1000,
        "amount_non_invoice": 0,
        "error_message": null
      }
    ]
    """
)

def process_receipts(pdf_path, template_path, output_path):
    print(f"🔄 {pdf_path} を解析中... (使用モデル: {MODEL_NAME})")

    # PDFファイルのアップロードと解析
    try:
        sample_file = genai.upload_file(path=pdf_path, display_name="Receipt PDF")
        
        print("   ☁️ Googleサーバーでファイルを処理しています...")
        while sample_file.state.name == "PROCESSING":
            time.sleep(2)
            sample_file = genai.get_file(sample_file.name)

        if sample_file.state.name == "FAILED":
            raise ValueError("Google側でのファイル処理に失敗しました。")

        response = model.generate_content([sample_file, "このPDFの全ページのレシート情報を抽出してください。"])
        
        response_text = response.text
        receipt_data = json.loads(response_text)
        print(f"✅ 解析完了: {len(receipt_data)} 件のデータを抽出しました。")

    except Exception as e:
        print(f"❌ AI解析エラー: {e}")
        return

    # Excelへの書き込み
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active 

        # ▼▼▼ ここを修正しました (9行目固定) ▼▼▼
        start_row = 9
        # ▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲▲

        for i, item in enumerate(receipt_data):
            row_num = start_row + i
            
            # --- マッピング処理 ---
            # B列: 支払日
            if item.get("date"):
                ws.cell(row=row_num, column=2).value = item["date"]
            
            # C列: 支払先
            if item.get("store_name"):
                ws.cell(row=row_num, column=3).value = item["store_name"]

            # 金額計算
            amt_8 = item.get("amount_8_percent") or 0
            amt_10 = item.get("amount_10_percent") or 0
            amt_other = item.get("amount_non_invoice") or 0

            # F列: 8%での支払い
            if amt_8 > 0:
                ws.cell(row=row_num, column=6).value = amt_8

            # G列: 8%以外の支払い
            total_other = amt_10 + amt_other
            if total_other > 0:
                ws.cell(row=row_num, column=7).value = total_other

            status_icon = "⚠️" if item.get("status") == "error" else "🆗"
            print(f"{status_icon} 行{row_num}: {item.get('date')} - {item.get('store_name')}")

        wb.save(output_path)
        print(f"🎉 完了しました！ファイル保存先: {output_path}")

    except Exception as e:
        print(f"❌ Excel保存エラー: {e}")

if __name__ == "__main__":
    INPUT_PDF = "scan_data.pdf"     
    TEMPLATE = "template.xlsx"      
    OUTPUT = "result_output.xlsx"   

    if os.path.exists(INPUT_PDF) and os.path.exists(TEMPLATE):
        process_receipts(INPUT_PDF, TEMPLATE, OUTPUT)
    else:
        print(f"エラー: {INPUT_PDF} または {TEMPLATE} が見つかりません。")